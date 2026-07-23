#!/usr/bin/env python3
"""
Vyplní evidenční Excel opatřeními z rizikových analýz.

Zdroje opatření:
  * Word RA (.docx)  – kapitola NAVRŽENÁ OPATŘENÍ (provozní i investiční)
  * PDF příloha (.pdf) – tabulka opatření jednoho druhu (provozní / investiční)

Do zvoleného listu se zapíší sloupce A–K:
  A Kalkulační jednice | B Typ objektu | C Název objektu | D Rok |
  E Druh opatření | F Číslo opatření | G Navržené opatření | H Zodpovídá |
  I Datum provedení | J Provedl | K Poznámka
Sloupce A a I–K zůstávají prázdné (doplňuje se ručně při plnění úkolů).
PDF přílohy nemají sloupec „Zodpovídá", proto u nich zůstane H prázdné.

Příklady:
  # Word RA -> list
  python ra_to_excel.py --template T.xlsx --out OUT.xlsx \
      --add "Morasice.docx::Morašice"

  # PDF přílohy (každá jeden druh) do jednoho listu
  python ra_to_excel.py --template T.xlsx --out OUT.xlsx \
      --pdf "Policsko_provozni.pdf::SV Polička::provozní" \
      --pdf "Policsko_investicni.pdf::SV Polička::investiční"

Ke stejnému listu lze přiřadit víc zdrojů; opatření se seřadí za sebou.
Pokud list v šabloně neexistuje, vytvoří se nový s hlavičkou.
"""
import argparse
from collections import OrderedDict
import openpyxl
from openpyxl.styles import Alignment, Font
from ra_extract import extract
from pdf_extract import extract_pdf

HEADER = ['Kalkulační jednice', 'Typ objektu', 'Název objektu', 'Rok',
          'Druh opatření', 'Číslo opatření', 'Navržené opatření', 'Zodpovídá',
          'Datum provedení', 'Provedl', 'Poznámka']
WIDTHS = {'A': 14, 'B': 12, 'C': 22, 'D': 7, 'E': 12, 'F': 11,
          'G': 60, 'H': 22, 'I': 12, 'J': 12, 'K': 22}


def write_sheet(ws, measures):
    if ws.max_row == 1 and all(c.value is None for c in ws[1]):
        for i, h in enumerate(HEADER, start=1):
            ws.cell(row=1, column=i, value=h).font = Font(bold=True)
    r = 2
    for m in measures:
        vals = ['', m['typ'], m['nazev'], m['year'], m['druh'],
                m['num'], m['measure'], m.get('resp', ''), '', '', '']
        for ci, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=ci, value=(v if v != '' else None))
            c.alignment = Alignment(wrap_text=True, vertical='top')
        r += 1
    for col, w in WIDTHS.items():
        ws.column_dimensions[col].width = w


def main():
    ap = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument('--template', help='vstupní .xlsx šablona (nepovinné)')
    ap.add_argument('--out', required=True, help='výstupní .xlsx')
    ap.add_argument('--add', action='append', default=[], metavar='DOCX::LIST',
                    help='Word RA a název listu (lze víckrát)')
    ap.add_argument('--pdf', action='append', default=[], metavar='PDF::LIST::DRUH',
                    help='PDF příloha, list a druh (provozní/investiční)')
    args = ap.parse_args()

    # posbírej opatření podle listu, v pořadí zadání
    per_sheet = OrderedDict()
    for spec in args.add:
        parts = spec.split('::')
        if len(parts) != 2:
            raise SystemExit(f"Chybný --add (očekávám DOCX::LIST): {spec}")
        path, sheet = parts
        per_sheet.setdefault(sheet, []).extend(extract(path))
    for spec in args.pdf:
        parts = spec.split('::')
        if len(parts) != 3:
            raise SystemExit(f"Chybný --pdf (očekávám PDF::LIST::DRUH): {spec}")
        path, sheet, druh = parts
        per_sheet.setdefault(sheet, []).extend(extract_pdf(path, druh))

    wb = openpyxl.load_workbook(args.template) if args.template else openpyxl.Workbook()
    if not args.template:
        wb.remove(wb.active)

    print(f"{'List':22} {'celkem':>7} {'prov':>5} {'inv':>5}")
    for sheet, ms in per_sheet.items():
        ws = wb[sheet] if sheet in wb.sheetnames else wb.create_sheet(sheet)
        write_sheet(ws, ms)
        pr = sum(1 for m in ms if m['druh'] == 'provozní')
        inv = sum(1 for m in ms if m['druh'] == 'investiční')
        print(f"{sheet:22} {len(ms):7} {pr:5} {inv:5}")

    wb.save(args.out)
    print(f"\nUloženo: {args.out}")


if __name__ == '__main__':
    main()
